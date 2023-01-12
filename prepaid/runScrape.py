from sys import excepthook
from selenium.webdriver import Edge
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By 
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.keys import Keys
import time
import os
import re
import openpyxl
from openpyxl import load_workbook   

#GLOBAL VAR---------------------------------------------------------------------------
SLEEPTIME = 20
SITE = "https://bianalytics.rci.rogers.com/t/BI/views/DailyPrepaidReportNewSubscriberSource/Result?:embed=yes#2"
dash = "-------------------"
DATE = ""
PROVINCE_NAME = ""
BRAND_NAME = ""
BRAND_DICT = dict([
    (0,"Chatr"),
    (1,"Fido"),
    (2,"Rogers")
])



PROVINCE_DICT = dict([
    (0,"BC"),
    (1,"AB"),
    (2,"SK"),
    (3,"MB"),
    (4,"ON"),
    (5,"PQ"),
    (6,"NB"),
    (7,"NS"),
    (8,"PE"),
    (9,"NL")
])

COUNT = 2
DOWNLOAD_DIRECTORY = ""
#ROOT = "C:\\Users\\Steven.Shan\\mypython\\tableau\\web scrape\\"
#ROOT = "C:\\Users\\Steven.Shan\\mypython\\tableau\\prepaid\\"
ROOT = "C:\\Users\\Wei.Li1\\Documents\\Prepaid\\"


#--------------------------------------------------------------------------------------------------------------------------
#initialize webdriver
options = webdriver.EdgeOptions()
options.use_chromium = True
options.add_argument("--start-maximized")
# options.add_argument("--headless")
# options.add_argument("disable-gpu")

#options.add_argument("--log-level=2")


#helpers:

#check whether the xpath exist in current webpage
def check_exists_by_xpath(xpath,web):
    try:
        web.find_element(By.XPATH, xpath)
    except NoSuchElementException:
        return False
    return True

#download and add downloaded info to the map excel sheet
def download(web):
    global PROVINCE_NAME
    global COUNT
    time.sleep(SLEEPTIME)
    el = WebDriverWait(web,300).until(EC.presence_of_element_located((By.XPATH, '//*[@id="download-ToolbarButton"]')))
    el.click()
    time.sleep(SLEEPTIME)
    el = WebDriverWait(web,300).until(EC.presence_of_element_located((By.XPATH, '//*[@id="DownloadDialog-Dialog-Body-Id"]/div/fieldset/button[3]')))
    el.click()
    time.sleep(SLEEPTIME)
    download = web.find_element(By.XPATH, '//*[@id="export-crosstab-options-dialog-Dialog-BodyWrapper-Dialog-Body-Id"]/div/div[3]/button')
    # download = WebDriverWait(web,300).until(EC.presence_of_element_located((By.XPATH, '//*[@id="export-crosstab-options-dialog-Dialog-BodyWrapper-Dialog-Body-Id"]/div/div[3]/button')))
    download.click()
    time.sleep(SLEEPTIME)
    
    os.rename(DOWNLOAD_DIRECTORY+"\\"+"Closing.xlsx",DOWNLOAD_DIRECTORY+"\\"+PROVINCE_NAME+"_"+BRAND_NAME+".xlsx")
    
    print(dash+"creating map"+dash) 

    wb = load_workbook(filename= DOWNLOAD_DIRECTORY+"\\MAP.xlsx")
    sheet = wb.active
    sheet["A"+str(COUNT)] = COUNT-1
    sheet["B"+str(COUNT)] = PROVINCE_NAME+"_"+BRAND_NAME
    sheet["C"+str(COUNT)] = PROVINCE_NAME
    sheet["D"+str(COUNT)] = BRAND_NAME
    wb.save(filename = DOWNLOAD_DIRECTORY+"\\MAP.xlsx")

    COUNT = COUNT + 1

    

#start SCRAPE!----------------------------------------------------------------------------------------------
#--------------------------------------------------------------------------------------------------------------------------        
#expand Province filter
def expand_Province_filter(web):
    el = WebDriverWait(web,30).until(EC.presence_of_element_located((By.XPATH, '//*[@id="tableau_base_widget_LegacyCategoricalQuickFilter_0"]/div/div[3]/span')))
    el.click()
    time.sleep(SLEEPTIME)
#click segment all
def click_Province_all(web):
    el = WebDriverWait(web,30).until(EC.presence_of_element_located((By.XPATH, '//*[@id="FI_federated.1xvzumc0s111ha1elqnms0pg4fie,none:PROVINCE:nk1292406501542173968_13281967033813092908_(All)"]/div[2]/input')))
    el.click()
    time.sleep(SLEEPTIME)

#traverse brand section and call traverse province for each brand
#all the brands will always exist
def traverse_Province(web):
    global DATE
    global PROVINCE_NAME
    global BRAND_NAME
    global DOWNLOAD_DIRECTORY
    time.sleep(30)
    PROVINCE_NAME = "ALL"
    print(dash+"Scraping"+DATE+"_"+PROVINCE_NAME+"_"+BRAND_NAME+dash)
    download(web)
    time.sleep(SLEEPTIME)
    j=0
    while j < 10:
        PROVINCE_NAME = PROVINCE_DICT[j]
        print(dash+"Scraping"+DATE+"_"+PROVINCE_NAME+"_"+BRAND_NAME+dash)
        expand_Province_filter(web)
        if j>0:
            click_Province_all(web)
        click_Province_all(web)
        el = WebDriverWait(web,50).until(EC.presence_of_element_located((By.XPATH, '//*[@id="FI_federated.1xvzumc0s111ha1elqnms0pg4fie,none:PROVINCE:nk1292406501542173968_13281967033813092908_'+str(j)+'"]/div[2]/input')))
        el.click()
        time.sleep(SLEEPTIME)
        el = WebDriverWait(web,30).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[6]')))
        el.click()
        time.sleep(SLEEPTIME)
        download(web)
        j+=1
    #get back to all
    expand_Province_filter(web)
    #select all
    click_Province_all(web)
    el = WebDriverWait(web,30).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[6]')))
    el.click()
    time.sleep(SLEEPTIME)


#--------------------------------------------------------------------------------------------------------------
#expand brand filter
def expand_brand_filter(web):
    el = WebDriverWait(web,30).until(EC.presence_of_element_located((By.XPATH, '//*[@id="tableau_base_widget_LegacyCategoricalQuickFilter_2"]/div/div[3]/span')))
    el.click()
    time.sleep(SLEEPTIME)
#click brand all button
def click_brand_all(web):
    el = WebDriverWait(web,50).until(EC.presence_of_element_located((By.XPATH, '//*[@id="FI_federated.1xvzumc0s111ha1elqnms0pg4fie,Brand1292406501542173968_13281967033813092908_(All)"]/div[2]/input')))
    el.click()
    time.sleep(SLEEPTIME)


#traverse segment section will call traverse brand for each segment
def traverse_brand(): 
    global DATE
    global PROVINCE_NAME
    global BRAND_NAME
    global DOWNLOAD_DIRECTORY
    time.sleep(SLEEPTIME)
    #print(dash+"Scraping Segment all")
    web = Edge(options=options)
    web.get(SITE) 
    time.sleep(SLEEPTIME)
    print(dash+"Scraping"+DATE+dash)
    time.sleep(SLEEPTIME)
    BRAND_NAME = "ALL"
    print(dash+"Scraping"+DATE+"_"+BRAND_NAME+dash)
    traverse_Province(web)
    k=0
    while k<3:
        web.quit() 
        web = Edge(options=options)
        web.get(SITE)
        time.sleep(SLEEPTIME)                                                                              
    
        print(dash+"Scraping"+DATE+dash)
        expand_brand_filter(web)
        click_brand_all(web)
        el = WebDriverWait(web,20).until(EC.presence_of_element_located((By.XPATH, '//*[@id="FI_federated.1xvzumc0s111ha1elqnms0pg4fie,Brand1292406501542173968_13281967033813092908_'+str(k)+'"]/div[2]/input')))
        el.click()
        time.sleep(SLEEPTIME)
        BRAND_NAME = BRAND_DICT[k]
        el = WebDriverWait(web,30).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[6]')))
        el.click()
        time.sleep(SLEEPTIME)
        print(dash+"Scraping"+DATE+"_"+BRAND_NAME+dash)
        traverse_Province(web)
        k+=1

    #get back to all
    expand_brand_filter(web)
    #select all
    click_brand_all(web)
    el = WebDriverWait(web,20).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[6]')))
    el.click()
    time.sleep(SLEEPTIME)
    print(dash+DATE+" is successfully scrapped"+dash)




#---------------------------------------------------------------------------------------------
#user interface


   
DATE = "Current_Month"
COUNT = 2
DOWNLOAD_DIRECTORY = ROOT+"Current_Month"
os.makedirs(DOWNLOAD_DIRECTORY)
print(dash+"folder should be created"+dash)    
print(dash+"creating map"+dash)
wb = openpyxl.Workbook()
wb.save(DOWNLOAD_DIRECTORY+"\\MAP.xlsx") 
wb = load_workbook(filename= DOWNLOAD_DIRECTORY+"\\MAP.xlsx")
sheet = wb.active
sheet["B1"] = "Selection#"
sheet["C1"] = "Province"
sheet["D1"] = "Brand"
wb.save(filename = DOWNLOAD_DIRECTORY+"\\MAP.xlsx")
options.add_experimental_option("prefs", {
    "download.default_directory": DOWNLOAD_DIRECTORY
})
traverse_brand()
print(dash+"successfully_scrapped   " +str(DATE)+"   "+dash)



print(dash+"Have A Great Day!"+dash)
        