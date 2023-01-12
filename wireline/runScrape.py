from sys import excepthook
from selenium.webdriver import Edge
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By 
from selenium.common.exceptions import NoSuchElementException
import time
import os
import re
import openpyxl
from openpyxl import load_workbook  
import sys 

#GLOBAL VAR---------------------------------------------------------------------------
SLEEPTIME = 10
dash = "-------------------"
DATE = ""
SEGMENT_NAME = ""
PROVINCE_NAME = ""
BRAND_NAME = ""
BRAND_DICT = dict([
    (0,"Legacy"),
    (1,"Ignite"),
    (2,"Fido"),
    (3,"TPIA")
])

SEGMENT_DICT = dict([
    (0,"Connected_Home"),
    (1,"Bulk"),
    (2,"Business"),
    (3,"Wholesale"),
    (4,"Business_Solution")
])
COUNT = 2
DOWNLOAD_DIRECTORY = ""
ROOT = ""
# ROOT = "C:\\Users\\Wei.Li1\\Documents\\Daily Cable TSU\\"

#--------------------------------------------------------------------------------------------------------------------------
#initialize webdriver
options = webdriver.EdgeOptions()
options.use_chromium = True
options.add_argument("--start-maximized")
# options.add_argument("--headless")
# options.add_argument("disable-gpu")
#options.add_argument("--log-level=2")



#helpers:



#check whether the xpath exist in current webpage
def check_exists_by_css(css,web):
    try:
        web.find_element(By.CSS_SELECTOR,css)
        #WebDriverWait(web,100).until(EC.presence_of_element_located((By.CSS_SELECTOR, css)))
    except NoSuchElementException:
        return False
    return True


#click by css
def click_B(css,web):
    try:
        el = WebDriverWait(web,300).until(EC.presence_of_element_located((By.CSS_SELECTOR, css)))
        el.click()
    except:
        time.sleep(SLEEPTIME)
        try:
            el = WebDriverWait(web,300).until(EC.presence_of_element_located((By.CSS_SELECTOR, css)))
            el.click()
        except:
            time.sleep(SLEEPTIME)
            try:
                el = WebDriverWait(web,300).until(EC.presence_of_element_located((By.CSS_SELECTOR, css)))
                el.click()
            except:
                time.sleep(SLEEPTIME)
                el = WebDriverWait(web,300).until(EC.presence_of_element_located((By.CSS_SELECTOR, css)))
                el.click()



#click by xpath
def click_x(xpath,web):
    try:
        el = WebDriverWait(web,300).until(EC.presence_of_element_located((By.XPATH, xpath)))
        el.click()
    except:
        time.sleep(SLEEPTIME)
        try:
            el = WebDriverWait(web,300).until(EC.presence_of_element_located((By.XPATH, xpath)))
            el.click()
        except:
            time.sleep(SLEEPTIME)
            try:
                el = WebDriverWait(web,300).until(EC.presence_of_element_located((By.XPATH, xpath)))
                el.click()
            except:
                time.sleep(SLEEPTIME)
                el = WebDriverWait(web,300).until(EC.presence_of_element_located((By.ID, 'loadingGlassPane')))
                el.click()
                time.sleep(SLEEPTIME)
                el = WebDriverWait(web,300).until(EC.presence_of_element_located((By.XPATH, xpath)))
                el.click()
        

#download and add downloaded info to the map excel sheet
def download(web):
    global PROVINCE_NAME
    global COUNT
    time.sleep(SLEEPTIME)
    click_B('#download-ToolbarButton',web)
    time.sleep(SLEEPTIME)
    click_B('#DownloadDialog-Dialog-Body-Id > div > fieldset > button:nth-child(4)',web)
    time.sleep(SLEEPTIME)
    click_B('#export-crosstab-options-dialog-Dialog-BodyWrapper-Dialog-Body-Id > div > div.fdr6v0d > button',web)
    #web.execute_script('arguments[0].click()', download)
    time.sleep(SLEEPTIME)
    month = str(re.match('^[^/]*', DATE).group(0))
    year = DATE[-4:]
    if PROVINCE_NAME == "British Columbia":
        PROVINCE_NAME = "BC"
    elif PROVINCE_NAME == "Northern Canada":
        PROVINCE_NAME = "NC"  
    time.sleep(SLEEPTIME)
    try:
        os.rename(DOWNLOAD_DIRECTORY+"\\"+"Totals.xlsx",DOWNLOAD_DIRECTORY+"\\"+month+"_"+year+"_"+SEGMENT_NAME+"_"+PROVINCE_NAME+"_"+BRAND_NAME+".xlsx")
    except:
        pass

    print(dash+"creating map"+dash) 

    wb = load_workbook(filename= DOWNLOAD_DIRECTORY+"\\MAP.xlsx")
    sheet = wb.active
    sheet["A"+str(COUNT)] = COUNT-1
    sheet["B"+str(COUNT)] = month+"_"+year+"_"+SEGMENT_NAME+"_"+PROVINCE_NAME+"_"+BRAND_NAME
    sheet["C"+str(COUNT)] = SEGMENT_NAME
    sheet["D"+str(COUNT)] = PROVINCE_NAME
    sheet["E"+str(COUNT)] = BRAND_NAME
    wb.save(filename = DOWNLOAD_DIRECTORY+"\\MAP.xlsx")

    COUNT = COUNT + 1

    

#start SCRAPE!----------------------------------------------------------------------------------------------

#expand province filter
def expand_province_filter(web):
    click_B('#tableau_base_widget_LegacyCategoricalQuickFilter_1 > div > div.CFContent > span',web)
    time.sleep(SLEEPTIME)

#click province all button
def click_province_all(web):
    click_B('#FI_federated\.07uceah05f3pzg12vx7ye10ec8da\,none\:PROVINCE_NM\:nk1863049714700099538_14391167526842317961_\(All\) > div.facetOverflow > input',web)
    time.sleep(SLEEPTIME)

#traverse the province section and download
#special cases: 
# 1. might not exist any province, download button does not exist
# 2. only one province exist, will influence how clicking convention work
def traverse_Province(web):
    global DATE
    global SEGMENT_NAME
    global PROVINCE_NAME
    global BRAND_NAME
    global DOWNLOAD_DIRECTORY
    time.sleep(SLEEPTIME)
    expand_province_filter(web)

    if check_exists_by_css('#FI_federated\.07uceah05f3pzg12vx7ye10ec8da\,none\:PROVINCE_NM\:nk1863049714700099538_14391167526842317961_0 > div.facetOverflow > input',web):
        if check_exists_by_css('#FI_federated\.07uceah05f3pzg12vx7ye10ec8da\,none\:PROVINCE_NM\:nk1863049714700099538_14391167526842317961_1 > div.facetOverflow > input' , web):
            click_x('/html/body/div[5]',web)
            time.sleep(SLEEPTIME)
            PROVINCE_NAME = "ALL"
            print(dash+"DOWNLOADING"+DATE+"_"+SEGMENT_NAME+BRAND_NAME+PROVINCE_NAME+dash)
            download(web)
            
            i=0
            while i<8:
                expand_province_filter(web)
                if check_exists_by_css('#FI_federated\.07uceah05f3pzg12vx7ye10ec8da\,none\:PROVINCE_NM\:nk1863049714700099538_14391167526842317961_'+str(i)+' > div.facetOverflow > input' ,web):
                    #if i==0 no need to click twice
                    if i>0:
                        click_province_all(web)
                    #for all click at least once to reset list to empty
                    click_province_all(web)
                    click_B('#FI_federated\.07uceah05f3pzg12vx7ye10ec8da\,none\:PROVINCE_NM\:nk1863049714700099538_14391167526842317961_'+str(i)+' > div.facetOverflow > input',web)
                    time.sleep(SLEEPTIME)
                    el = WebDriverWait(web,300).until(EC.presence_of_element_located((By.CSS_SELECTOR, '#FI_federated\.07uceah05f3pzg12vx7ye10ec8da\,none\:PROVINCE_NM\:nk1863049714700099538_14391167526842317961_'+str(i)+' > div.facetOverflow > a')))
                    PROVINCE_NAME = el.get_attribute("title")
                    time.sleep(SLEEPTIME)
                    click_x('/html/body/div[5]',web)
                    # web.refresh()
                    time.sleep(SLEEPTIME)
                    print(dash+"DOWNLOADING"+DATE+"_"+SEGMENT_NAME+"_"+BRAND_NAME+"_"+PROVINCE_NAME+dash)
                    download(web)
                    i+=1
                else:
                    click_x('/html/body/div[5]',web)
                    # web.refresh()
                    time.sleep(SLEEPTIME)
                    break
            #get back to all
            expand_province_filter(web)
            #select all
            click_province_all(web)
            click_x('/html/body/div[5]',web)
            time.sleep(SLEEPTIME)
        #only one sheet is available to download
        else:
            print(dash+"only one sheet is available(special case)"+dash)
            el = WebDriverWait(web,500).until(EC.presence_of_element_located((By.CSS_SELECTOR, '#FI_federated\.07uceah05f3pzg12vx7ye10ec8da\,none\:PROVINCE_NM\:nk1863049714700099538_14391167526842317961_0 > div.facetOverflow > a')))
            PROVINCE_NAME = el.get_attribute("title")
            print(dash+"DOWNLOADING"+DATE+"_"+SEGMENT_NAME+"_"+BRAND_NAME+"_"+PROVINCE_NAME+dash)
            click_x('/html/body/div[5]',web)
            time.sleep(SLEEPTIME)
            download(web)
    #no province is available
    else:
        print(dash+"nothing to download"+dash)
        click_x('/html/body/div[5]',web)
        time.sleep(SLEEPTIME)


#--------------------------------------------------------------------------------------------------------------------------        

#expand brand filter
def expand_brand_filter(web):
    click_B('#tableau_base_widget_LegacyCategoricalQuickFilter_3 > div > div.CFContent > span',web)
    time.sleep(SLEEPTIME)
#click brand all button
def click_brand_all(web):
    click_B('#FI_federated\.07uceah05f3pzg12vx7ye10ec8da\,none\:PRODUCT_BRAND\:nk1863049714700099538_14391167526842317961_\(All\) > div.facetOverflow > input',web)
    time.sleep(SLEEPTIME)
#traverse brand section and call traverse province for each brand
#all the brands will always exist
def traverse_brand(web):
    global DATE
    global SEGMENT_NAME
    global PROVINCE_NAME
    global BRAND_NAME
    global DOWNLOAD_DIRECTORY
    time.sleep(SLEEPTIME)
    BRAND_NAME = "ALL"
    print(dash+"Scraping"+DATE+"_"+SEGMENT_NAME+"_"+BRAND_NAME+dash)
    traverse_Province(web)
    j=0
    while j < 4:
        BRAND_NAME = BRAND_DICT[j]
        print(dash+"Scraping"+DATE+"_"+SEGMENT_NAME+"_"+BRAND_NAME+dash)
        expand_brand_filter(web)
        if j>0:
            click_brand_all(web)
        click_brand_all(web)
        click_B('#FI_federated\.07uceah05f3pzg12vx7ye10ec8da\,none\:PRODUCT_BRAND\:nk1863049714700099538_14391167526842317961_'+str(j)+' > div.facetOverflow > input',web)
        time.sleep(SLEEPTIME)
        click_x('/html/body/div[5]',web)
        time.sleep(SLEEPTIME)
        traverse_Province(web)
        j+=1
    #get back to all
    expand_brand_filter(web)
    #select all
    click_brand_all(web)
    click_x('/html/body/div[5]',web)
    
    time.sleep(SLEEPTIME)


#--------------------------------------------------------------------------------------------------------------
#expand date filter
def expand_date_selector(web):
    click_B('#tableau_base_widget_LegacyCategoricalQuickFilter_5 > div > div.CFContent > span',web)
    #el.click()
    time.sleep(SLEEPTIME)
#click date all button
def click_date_all(web):
    click_B('#FI_federated\.07uceah05f3pzg12vx7ye10ec8da\,none\:Calculation_783907863373090816\:ok1863049714700099538_14391167526842317961_\(All\) > div.facetOverflow > input',web)
    time.sleep(SLEEPTIME)
#expand segment filter
def expand_Segment_filter(web):
    click_B('#tableau_base_widget_LegacyCategoricalQuickFilter_2 > div > div.CFContent > span',web)
    time.sleep(SLEEPTIME)
#click segment all
def click_Segment_all(web):
    click_B('#FI_federated\.07uceah05f3pzg12vx7ye10ec8da\,none\:SEGMENT_NM\:nk1863049714700099538_14391167526842317961_\(All\) > div.facetOverflow > input',web)
    time.sleep(SLEEPTIME)
#traverse segment section will call traverse brand for each segment
def traverse_Segment(input): 
    global DATE
    global SEGMENT_NAME
    global PROVINCE_NAME
    global BRAND_NAME
    global DOWNLOAD_DIRECTORY
    
    time.sleep(SLEEPTIME)
    #print(dash+"Scraping Segment all")
    web = Edge(options=options)
    web.get("https://bianalytics.rci.rogers.com/t/BI/views/DailyCableTSU/DailyCableTSU?:embed=yes&:display_count=no&:showVizHome=no&:origin=viz_share_link#1") 
    time.sleep(SLEEPTIME)
    expand_date_selector(web)
    click_date_all(web)
    click_date_all(web)
    click_B('#FI_federated\.07uceah05f3pzg12vx7ye10ec8da\,none\:Calculation_783907863373090816\:ok1863049714700099538_14391167526842317961_'+str(input)+' > div.facetOverflow > input',web)
    
    time.sleep(SLEEPTIME)
    el = WebDriverWait(web,300).until(EC.presence_of_element_located((By.CSS_SELECTOR, '#FI_federated\.07uceah05f3pzg12vx7ye10ec8da\,none\:Calculation_783907863373090816\:ok1863049714700099538_14391167526842317961_'+str(input)+' > div.facetOverflow > a')))
    DATE =el.get_attribute("title")
    print(dash+"Scraping"+DATE+dash)
    
    
    click_x('/html/body/div[5]',web)
    time.sleep(SLEEPTIME)
    SEGMENT_NAME = "ALL"
    print(dash+"Scraping"+DATE+"_"+SEGMENT_NAME+dash)
    traverse_brand(web)
    
    k=0
    while k<5:
        web.quit() 
        web = Edge(options=options)
        web.get("https://bianalytics.rci.rogers.com/t/BI/views/DailyCableTSU/DailyCableTSU?:embed=yes&:display_count=no&:showVizHome=no&:origin=viz_share_link#1")
        time.sleep(SLEEPTIME)
        expand_date_selector(web)
        click_date_all(web)
        click_date_all(web)
        click_B('#FI_federated\.07uceah05f3pzg12vx7ye10ec8da\,none\:Calculation_783907863373090816\:ok1863049714700099538_14391167526842317961_'+str(input)+' > div.facetOverflow > input',web)
        time.sleep(SLEEPTIME)
        click_x('/html/body/div[5]',web)
        time.sleep(SLEEPTIME)
        print(dash+"Scraping"+DATE+dash)
        expand_Segment_filter(web)
        click_Segment_all(web)
        click_B('#FI_federated\.07uceah05f3pzg12vx7ye10ec8da\,none\:SEGMENT_NM\:nk1863049714700099538_14391167526842317961_'+str(k)+' > div.facetOverflow > input',web)
        time.sleep(SLEEPTIME)
        SEGMENT_NAME = SEGMENT_DICT[k]

        #get out of filter selection - click on webpage body
        click_x('/html/body/div[5]',web)
        time.sleep(SLEEPTIME)
        print(dash+"Scraping"+DATE+"_"+SEGMENT_NAME+dash)
        traverse_brand(web)
        k+=1
        

    try:
        #get back to all
        expand_Segment_filter(web)
        #select all
        click_Segment_all(web)
        el = WebDriverWait(web,300).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]')))
        el.click()
        time.sleep(SLEEPTIME)
        print(dash+DATE+" is successfully scrapped"+dash)
    except:
        print(dash+"warning"+dash)




#---------------------------------------------------------------------------------------------

ROOT = sys.argv[2]
input = sys.argv[1]

#row count for map excel
COUNT = 2   
web = Edge(options=options)
web.get("https://bianalytics.rci.rogers.com/t/BI/views/DailyCableTSU/DailyCableTSU?:embed=yes&:display_count=no&:showVizHome=no&:origin=viz_share_link#1") 
time.sleep(SLEEPTIME)
expand_date_selector(web)
print("date_selector expanded")
el = WebDriverWait(web,300).until(EC.presence_of_element_located((By.CSS_SELECTOR, '#FI_federated\.07uceah05f3pzg12vx7ye10ec8da\,none\:Calculation_783907863373090816\:ok1863049714700099538_14391167526842317961_'+str(input)+' > div.facetOverflow > a')))
mydate = el.get_attribute("title")
month = int(re.match('^[^/]*', mydate).group(0))
year = mydate[-4:]
print(dash+"month is: " +str(month)+" year is: "+year+dash)
web.quit()

DOWNLOAD_DIRECTORY = ROOT+year+"_"+str(month)
os.makedirs(DOWNLOAD_DIRECTORY)
print(dash+"folder should be created"+dash)    
print(dash+"creating map"+dash)
wb = openpyxl.Workbook()
wb.save(DOWNLOAD_DIRECTORY+"\\MAP.xlsx") 
wb = load_workbook(filename= DOWNLOAD_DIRECTORY+"\\MAP.xlsx")
sheet = wb.active
sheet["B1"] = "Selection#"
sheet["C1"] = "Segment"
sheet["D1"] = "Province"
sheet["E1"] = "Brand"
wb.save(filename = DOWNLOAD_DIRECTORY+"\\MAP.xlsx")
options.add_experimental_option("prefs", {
    "download.default_directory": DOWNLOAD_DIRECTORY
})
traverse_Segment(input)
print(dash+"successfully_scrapped   " +str(input)+"   "+dash)





print(dash+"Have A Great Day!"+dash)
        