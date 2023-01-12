from sys import excepthook
from selenium.webdriver import Edge
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By 
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.keys import Keys
import time
import os
import re
import openpyxl
from openpyxl import load_workbook   

#GLOBAL VAR---------------------------------------------------------------------------
SLEEPTIME = 20
SITE = "https://bianalytics.rci.rogers.com/t/BI/views/DailyPostpaidReportNewSubscriberSource/Postpaid?:embed=yes#1"
dash = "-------------------"
DATE = ""
SEGMENT_NAME = ""
PROVINCE_NAME = ""
BRAND_NAME = ""
BRAND_DICT = dict([
    (0,"Rogers"),
    (1,"Fido"),
    (2,"Cityfone")
])

SEGMENT_DICT = dict([
    (0,"Consumer"),
    (1,"Rogers_for_Business")
])

PROVINCE_DICT = dict([
    (0,"BC"),
    (1,"AB"),
    (2,"SK"),
    (3,"MB"),
    (4,"ON"),
    (5,"PQ"),
    (6,"NB"),
    (7,"NS"),
    (8,"PE"),
    (9,"NL")
])
MONTH_DICT_LEAP = dict([
    (1,"31"),
    (2,"29"),
    (3,"31"),
    (4,"30"),
    (5,"31"),
    (6,"30"),
    (7,"31"),
    (8,"31"),
    (9,"30"),
    (10,"31"),
    (11,"30"),
    (12,"31")
])
MONTH_DICT = dict([
    (1,"31"),
    (2,"28"),
    (3,"31"),
    (4,"30"),
    (5,"31"),
    (6,"30"),
    (7,"31"),
    (8,"31"),
    (9,"30"),
    (10,"31"),
    (11,"30"),
    (12,"31")
])
COUNT = 2
DOWNLOAD_DIRECTORY = ""
#ROOT = "C:\\Users\\Steven.Shan\\mypython\\tableau\\web scrape\\"
ROOT = "C:\\Users\\Steven.Shan\\mypython\\tableau\\postpaid\\"
#ROOT = "C:\\Users\\Wei.Li1\\Documents\\Postpaid\\"

#--------------------------------------------------------------------------------------------------------------------------
#initialize webdriver
options = webdriver.EdgeOptions()
options.use_chromium = True
options.add_argument("--start-maximized")
# options.add_argument("--headless")
# options.add_argument("disable-gpu")

#options.add_argument("--log-level=2")


#helpers:

#check whether the xpath exist in current webpage
def check_exists_by_xpath(xpath,web):
    try:
        web.find_element(By.XPATH, xpath)
    except NoSuchElementException:
        return False
    return True

#download and add downloaded info to the map excel sheet
def download(web):
    global PROVINCE_NAME
    global COUNT
    time.sleep(SLEEPTIME)
    el = WebDriverWait(web,300).until(EC.presence_of_element_located((By.XPATH, '//*[@id="download-ToolbarButton"]')))
    el.click()
    time.sleep(SLEEPTIME)
    el = WebDriverWait(web,300).until(EC.presence_of_element_located((By.XPATH, '//*[@id="DownloadDialog-Dialog-Body-Id"]/div/fieldset/button[3]')))
    el.click()
    time.sleep(SLEEPTIME)
    el = WebDriverWait(web,300).until(EC.presence_of_element_located((By.XPATH, '//*[@id="export-crosstab-options-dialog-Dialog-BodyWrapper-Dialog-Body-Id"]/div/div[1]/div[2]/div/div/div[2]/div')))
    el.click()
    time.sleep(SLEEPTIME)
    download = web.find_element(By.XPATH, '//*[@id="export-crosstab-options-dialog-Dialog-BodyWrapper-Dialog-Body-Id"]/div/div[3]/button')
    # download = WebDriverWait(web,300).until(EC.presence_of_element_located((By.XPATH, '//*[@id="export-crosstab-options-dialog-Dialog-BodyWrapper-Dialog-Body-Id"]/div/div[3]/button')))
    download.click()
    time.sleep(SLEEPTIME)
    month = str(re.match('^[^/]*', DATE).group(0))
    year = DATE[-4:]


    try:
        
        os.rename(DOWNLOAD_DIRECTORY+"\\"+"Closing SC.xlsx",DOWNLOAD_DIRECTORY+"\\"+month+"_"+year+"_"+SEGMENT_NAME+"_"+PROVINCE_NAME+"_"+BRAND_NAME+".xlsx")
        
        print(dash+"creating map"+dash) 

        wb = load_workbook(filename= DOWNLOAD_DIRECTORY+"\\MAP.xlsx")
        sheet = wb.active
        sheet["A"+str(COUNT)] = COUNT-1
        sheet["B"+str(COUNT)] = month+"_"+year+"_"+SEGMENT_NAME+"_"+PROVINCE_NAME+"_"+BRAND_NAME
        sheet["C"+str(COUNT)] = SEGMENT_NAME
        sheet["D"+str(COUNT)] = PROVINCE_NAME
        sheet["E"+str(COUNT)] = BRAND_NAME
        wb.save(filename = DOWNLOAD_DIRECTORY+"\\MAP.xlsx")

        COUNT = COUNT + 1
    except:
        COUNT=COUNT

    

#start SCRAPE!----------------------------------------------------------------------------------------------

#expand segment filter
def expand_Segment_filter(web):
    el = WebDriverWait(web,30).until(EC.presence_of_element_located((By.XPATH, '//*[@id="tableau_base_widget_LegacyCategoricalQuickFilter_1"]/div/div[3]/span')))
    el.click()
    time.sleep(SLEEPTIME)
#click segment all
def click_Segment_all(web):
    el = WebDriverWait(web,30).until(EC.presence_of_element_located((By.XPATH, '//*[@id="FI_federated.0dn4toz0kzraq91gkruzo1k7pvrl,Segment_113277616459342307234_834514642350544244_(All)"]/div[2]/input')))
    el.click()
    time.sleep(SLEEPTIME)

#traverse the province section and download
#special cases: 
# 1. might not exist any province, download button does not exist
# 2. only one province exist, will influence how clicking convention work
def traverse_Segment(web):
    global DATE
    global SEGMENT_NAME
    global PROVINCE_NAME
    global BRAND_NAME
    global DOWNLOAD_DIRECTORY
    time.sleep(SLEEPTIME)
    expand_Segment_filter(web)
    if check_exists_by_xpath('//*[@id="FI_federated.0dn4toz0kzraq91gkruzo1k7pvrl,Segment_113277616459342307234_834514642350544244_0"]/div[2]/input',web):
        if check_exists_by_xpath('//*[@id="FI_federated.0dn4toz0kzraq91gkruzo1k7pvrl,Segment_113277616459342307234_834514642350544244_1"]/div[2]/input' , web):
            el = WebDriverWait(web,30).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]')))
            el.click()
            time.sleep(SLEEPTIME)
            SEGMENT_NAME = "ALL"
            print(dash+"DOWNLOADING"+DATE+"_"+SEGMENT_NAME+BRAND_NAME+PROVINCE_NAME+dash)
            download(web)
            i=0
            while i<2:
                expand_Segment_filter(web)
                if check_exists_by_xpath('//*[@id="FI_federated.0dn4toz0kzraq91gkruzo1k7pvrl,Segment_113277616459342307234_834514642350544244_'+str(i)+'"]/div[2]/input' ,web):
                    #if i==0 no need to click twice
                    if i>0:
                        click_Segment_all(web)
                    #for all click at least once to reset list to empty
                    click_Segment_all(web)
                    el = WebDriverWait(web,30).until(EC.presence_of_element_located((By.XPATH, '//*[@id="FI_federated.0dn4toz0kzraq91gkruzo1k7pvrl,Segment_113277616459342307234_834514642350544244_'+str(i)+'"]/div[2]/input')))
                    el.click()
                    time.sleep(SLEEPTIME)
                    SEGMENT_NAME = web.find_element(By.XPATH, '//*[@id="FI_federated.0dn4toz0kzraq91gkruzo1k7pvrl,Segment_113277616459342307234_834514642350544244_'+str(i)+'"]/div[2]/a').get_attribute("title")
                    time.sleep(SLEEPTIME)
                    el = WebDriverWait(web,30).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]')))
                    el.click()
                    # web.refresh()
                    time.sleep(SLEEPTIME)
                    print(dash+"DOWNLOADING"+DATE+"_"+SEGMENT_NAME+"_"+BRAND_NAME+"_"+PROVINCE_NAME+dash)
                    download(web)
                    i+=1
                else:
                    el = WebDriverWait(web,30).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]')))
                    el.click()
                    # web.refresh()
                    time.sleep(SLEEPTIME)
                    break
            #get back to all
            expand_Segment_filter(web)
            #select all
            click_Segment_all(web)
            el = WebDriverWait(web,50).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]')))
            el.click()
            time.sleep(SLEEPTIME)
        else:
            print(dash+"only one sheet is available(special case)"+dash)
            SEGMENT_NAME = web.find_element(By.XPATH, '//*[@id="FI_federated.07uceah05f3pzg12vx7ye10ec8da,none:PROVINCE_NM:nk1863049714700099538_14391167526842317961_0"]/div[2]/a').get_attribute("title")
            print(dash+"DOWNLOADING"+DATE+"_"+SEGMENT_NAME+"_"+BRAND_NAME+"_"+PROVINCE_NAME+dash)
            el = WebDriverWait(web,30).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]')))
            el.click()
            time.sleep(SLEEPTIME)
            download(web)
        
    else:
        print(dash+"nothing to download"+dash)
        el = WebDriverWait(web,30).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]')))
        el.click()
        time.sleep(SLEEPTIME)


#--------------------------------------------------------------------------------------------------------------------------        
#expand Province filter
def expand_Province_filter(web):
    el = WebDriverWait(web,30).until(EC.presence_of_element_located((By.XPATH, '//*[@id="tableau_base_widget_LegacyCategoricalQuickFilter_3"]/div/div[3]/span')))
    el.click()
    time.sleep(SLEEPTIME)
#click segment all
def click_Province_all(web):
    el = WebDriverWait(web,30).until(EC.presence_of_element_located((By.XPATH, '//*[@id="FI_federated.0dn4toz0kzraq91gkruzo1k7pvrl,none:PROVINCE:nk13277616459342307234_834514642350544244_(All)"]/div[2]/input')))
    el.click()
    time.sleep(SLEEPTIME)

#traverse brand section and call traverse province for each brand
#all the brands will always exist
def traverse_Province(web):
    global DATE
    global SEGMENT_NAME
    global PROVINCE_NAME
    global BRAND_NAME
    global DOWNLOAD_DIRECTORY
    time.sleep(SLEEPTIME)
    PROVINCE_NAME = "ALL"
    print(dash+"Scraping"+DATE+"_"+PROVINCE_NAME+"_"+BRAND_NAME+dash)
    traverse_Segment(web)
    j=0
    while j < 10:
        PROVINCE_NAME = PROVINCE_DICT[j]
        print(dash+"Scraping"+DATE+"_"+PROVINCE_NAME+"_"+BRAND_NAME+dash)
        expand_Province_filter(web)
        if j>0:
            click_Province_all(web)
        click_Province_all(web)
        el = WebDriverWait(web,50).until(EC.presence_of_element_located((By.XPATH, '//*[@id="FI_federated.0dn4toz0kzraq91gkruzo1k7pvrl,none:PROVINCE:nk13277616459342307234_834514642350544244_'+str(j)+'"]/div[2]/input')))
        el.click()
        time.sleep(SLEEPTIME)
        el = WebDriverWait(web,30).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]')))
        el.click()
        time.sleep(SLEEPTIME)
        traverse_Segment(web)
        j+=1
    #get back to all
    expand_Province_filter(web)
    #select all
    click_Province_all(web)
    el = WebDriverWait(web,30).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]')))
    el.click()
    time.sleep(SLEEPTIME)


#--------------------------------------------------------------------------------------------------------------
#type date in the box
def input_date(web):
    el = WebDriverWait(web,30).until(EC.presence_of_element_located((By.XPATH, '//*[@id="typein_[Parameters].[Report Date Parameter]"]/span[1]/input')))
    el.send_keys(Keys.CONTROL + "a")
    el.send_keys(Keys.DELETE)
    time.sleep(SLEEPTIME)
    el.send_keys(DATE)
    time.sleep(SLEEPTIME)
#expand brand filter
def expand_brand_filter(web):
    el = WebDriverWait(web,30).until(EC.presence_of_element_located((By.XPATH, '//*[@id="tableau_base_widget_LegacyCategoricalQuickFilter_5"]/div/div[3]/span')))
    el.click()
    time.sleep(SLEEPTIME)
#click brand all button
def click_brand_all(web):
    el = WebDriverWait(web,50).until(EC.presence_of_element_located((By.XPATH, '//*[@id="FI_federated.0dn4toz0kzraq91gkruzo1k7pvrl,Brand13277616459342307234_834514642350544244_(All)"]/div[2]/input')))
    el.click()
    time.sleep(SLEEPTIME)


#traverse segment section will call traverse brand for each segment
def traverse_brand(): 
    global DATE
    global SEGMENT_NAME
    global PROVINCE_NAME
    global BRAND_NAME
    global DOWNLOAD_DIRECTORY
    time.sleep(SLEEPTIME)
    #print(dash+"Scraping Segment all")
    web = Edge(options=options)
    web.get(SITE) 
    time.sleep(15)
    input_date(web)
    print(dash+"Scraping"+DATE+dash)
    time.sleep(SLEEPTIME)
    BRAND_NAME = "ALL"
    print(dash+"Scraping"+DATE+"_"+BRAND_NAME+dash)
    traverse_Province(web)
    k=0
    while k<3:
        web.quit() 
        web = Edge(options=options)
        web.get(SITE)
        time.sleep(15)
        input_date(web)
        print(dash+"Scraping"+DATE+dash)
        expand_brand_filter(web)
        click_brand_all(web)
        el = WebDriverWait(web,20).until(EC.presence_of_element_located((By.XPATH, '//*[@id="FI_federated.0dn4toz0kzraq91gkruzo1k7pvrl,Brand13277616459342307234_834514642350544244_'+str(k)+'"]/div[2]/input')))
        el.click()
        time.sleep(SLEEPTIME)
        BRAND_NAME = BRAND_DICT[k]
        el = WebDriverWait(web,30).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]')))
        el.click()
        time.sleep(SLEEPTIME)
        print(dash+"Scraping"+DATE+"_"+SEGMENT_NAME+dash)
        traverse_Province(web)
        k+=1

    #get back to all
    expand_brand_filter(web)
    #select all
    click_brand_all(web)
    el = WebDriverWait(web,20).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]')))
    el.click()
    time.sleep(SLEEPTIME)
    print(dash+DATE+" is successfully scrapped"+dash)




#---------------------------------------------------------------------------------------------
#user interface

print(dash+"Enter the year you want to download: "+dash)
year = str(input())
print(dash+"Enter the month you want to download or -1 to download the entire year:"+dash)
month = int(input())


if month != -1:
    if int(year)%4==0:
        DATE = str(month)+"/"+MONTH_DICT_LEAP[month]+"/"+year
    else:
        DATE = str(month)+"/"+MONTH_DICT[month]+"/"+year
    COUNT = 2
    DOWNLOAD_DIRECTORY = ROOT+year+"_"+str(month)
    os.makedirs(DOWNLOAD_DIRECTORY)
    print(dash+"folder should be created"+dash)    
    print(dash+"creating map"+dash)
    wb = openpyxl.Workbook()
    wb.save(DOWNLOAD_DIRECTORY+"\\MAP.xlsx") 
    wb = load_workbook(filename= DOWNLOAD_DIRECTORY+"\\MAP.xlsx")
    sheet = wb.active
    sheet["B1"] = "Selection#"
    sheet["C1"] = "Segment"
    sheet["D1"] = "Province"
    sheet["E1"] = "Brand"
    wb.save(filename = DOWNLOAD_DIRECTORY+"\\MAP.xlsx")
    options.add_experimental_option("prefs", {
        "download.default_directory": DOWNLOAD_DIRECTORY
    })
    traverse_brand()
    print(dash+"successfully_scrapped   " +str(DATE)+"   "+dash)
else:
    DOWNLOAD_DIRECTORY = ROOT+year
    os.makedirs(DOWNLOAD_DIRECTORY)
    print(dash+"folder should be created"+dash)    
    print(dash+"creating map"+dash)
    wb = openpyxl.Workbook()
    wb.save(DOWNLOAD_DIRECTORY+"\\MAP.xlsx") 
    wb = load_workbook(filename= DOWNLOAD_DIRECTORY+"\\MAP.xlsx")
    sheet = wb.active
    sheet["B1"] = "Selection#"
    sheet["C1"] = "Segment"
    sheet["D1"] = "Province"
    sheet["E1"] = "Brand"
    wb.save(filename = DOWNLOAD_DIRECTORY+"\\MAP.xlsx")
    options.add_experimental_option("prefs", {
        "download.default_directory": DOWNLOAD_DIRECTORY
    })
    COUNT = 2
    x = 1
    while x < 13:
        if year%4==0:
            DATE = str(month)+"/"+MONTH_DICT_LEAP[x]+"/"+year
        else:
            DATE = str(month)+"/"+MONTH_DICT[x]+"/"+year
        traverse_brand()
        print(dash+"successfully_scrapped   " +str(DATE)+"   "+dash)
        x=x+1


print(dash+"Have A Great Day!"+dash)
        